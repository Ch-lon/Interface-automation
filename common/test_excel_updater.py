# 自动更新测试文件，使其可以通过重复性校验，能够成功上传

import os
from pathlib import Path
import openpyxl
from db_handler import db
import settings


class ExcelUpdater:
    def __init__(self):
        self.test_file_dir = Path(settings.TEST_FILE_DIR).resolve()
        self.valid_sheets = [
            "新增明细", "修改明细", "删除明细",
            "新增学历", "修改学历", "删除学历", "人才现职"
        ]

    def execute_query(self, sql):
        try:
            return db.get_all(sql)
        except Exception as e:
            print(f"错误执行查询: {e}")
            return []

    def modify_value(self, value):
        if not value or not isinstance(value, str):
            return "1"
        if value[-1].isdigit():
            last_digit = int(value[-1])
            return value[:-1] + str(last_digit + 1)
        else:
            return value + "1"

    def _get_priority_fields(self, uni_check_list):
        priority_fields = ["ver_no", "elected_year"]
        return [f for f in priority_fields if f in uni_check_list] + \
               [f for f in uni_check_list if f not in priority_fields]

    def process_new_details_sheet(self, wb, sheet_name):
        print(f"处理 '{sheet_name}' 表单...")
        sheet = wb[sheet_name]
        vc_column = next((col for col in range(1, sheet.max_column + 1)
                          if sheet.cell(3, col).value == "var_code"), None)
        if not vc_column:
            print("  ⚠️ 跳过：未找到'var_code'列（第3行）")
            return

        modified_records = []
        for row in range(4, sheet.max_row + 1):
            var_code = sheet.cell(row, vc_column).value
            if not var_code:
                continue

            uni_fields = self._get_priority_fields([
                r[0] for r in self.execute_query(
                    f"select field_key from details.var_det_field "
                    f"where def_id = (select det_def_id from details.v_var where code = '{var_code}') "
                    f"and uni_check > 0"
                )
            ])

            for field in uni_fields:
                col = next((c for c in range(1, sheet.max_column + 1)
                           if sheet.cell(3, c).value == field), None)
                if not col:
                    continue

                original_value = sheet.cell(row, col).value
                if field in ["ver_no", "elected_year"] and isinstance(original_value, (int, float)):
                    new_value = int(original_value) + 1
                elif original_value:
                    new_value = self.modify_value(str(original_value))
                else:
                    continue

                if original_value == new_value:
                    continue

                sheet.cell(row, col).value = new_value
                modified_records.append((field, original_value, new_value, row))
                break

        for field, old_val, new_val, row in modified_records:
            print(f"  ✅ 修改字段：{field} | 原值: {old_val} → 新值: {new_val}（行{row}）")
        print(f"  ✔️ 完成：共处理{len(modified_records)}条唯一标识记录")

    def process_modify_details_sheet(self, wb, sheet_name):
        print(f"处理 '{sheet_name}' 表单...")
        sheet = wb[sheet_name]
        rid_col = next((c for r in [2, 3] for c in range(1, sheet.max_column + 1)
                        if sheet.cell(r, c).value == "rid"), None)
        update_fields_col = next((c for r in [2, 3] for c in range(1, sheet.max_column + 1)
                                  if sheet.cell(r, c).value == "update_fields"), None)
        if not rid_col or not update_fields_col:
            print(f"  ⚠️ 跳过：未找到'rid'或'update_fields'列（2-3行）")
            return

        modified_records = []
        for row in range(4, sheet.max_row + 1):
            rid = sheet.cell(row, rid_col).value
            update_fields = sheet.cell(row, update_fields_col).value
            if not rid or not update_fields:
                continue

            uni_fields = self._get_priority_fields([
                r[0] for r in self.execute_query(
                    f"select field_key from details.var_det_field "
                    f"where def_id = (select det_def_id from details.variable "
                    f"where code = (select var_code from details.v_var_detail_xp where rid = {rid})) "
                    f"and uni_check > 0"
                )
            ])

            fields_to_update = update_fields.split(",") if update_fields != "ALL_FIELDS" else uni_fields

            for field in uni_fields:
                if field not in fields_to_update:
                    continue

                col = next((c for c in range(1, sheet.max_column + 1)
                           if sheet.cell(3, c).value == field), None)
                if not col:
                    continue

                original_value = sheet.cell(row, col).value
                if field in ["ver_no", "elected_year"] and isinstance(original_value, (int, float)):
                    new_value = int(original_value) + 1
                elif original_value:
                    new_value = self.modify_value(str(original_value))
                else:
                    continue

                if original_value == new_value:
                    continue

                sheet.cell(row, col).value = new_value
                modified_records.append((field, original_value, new_value, row))
                break

        for field, old_val, new_val, row in modified_records:
            print(f"  ✅ 修改字段：{field} | 原值: {old_val} → 新值: {new_val}（行{row}）")
        print(f"  ✔️ 完成：共处理{len(modified_records)}条修改记录")

    def process_delete_details_sheet(self, wb, sheet_name):
        print(f"处理 '{sheet_name}' 表单...")
        sheet = wb[sheet_name]
        rid_col = next((col for col in range(1, sheet.max_column + 1)
                        if sheet.cell(1, col).value == "rid"), None)
        if not rid_col:
            print(f"  ⚠️ 跳过：未找到'rid'列（第1行）")
            return

        sql = f"select rid from details.v_var_detail_xp where var_code = 'VG0008' limit {sheet.max_row - 1}"
        new_rids = [str(rid[0]) for rid in self.execute_query(sql)]

        for i, row in enumerate(range(2, sheet.max_row + 1)):
            if i < len(new_rids):
                sheet.cell(row, rid_col).value = new_rids[i]
            else:
                sheet.delete_rows(row, sheet.max_row - row + 1)
                break
        print(f"  ✔️ 完成：更新{len(new_rids)}条rid记录")

    def process_talent_code_sheet(self, wb, sheet_name):
        print(f"处理 '{sheet_name}' 表单...")
        sheet = wb[sheet_name]
        talent_code_col = next(
            (col for row in [2, 3] for col in range(1, sheet.max_column + 1)
             if sheet.cell(row, col).value == "talent_code"),
            None
        )
        if not talent_code_col:
            print(f"  ⚠️ 跳过：未找到'talent_code'列（2-3行）")
            return

        modified_count = 0
        for row in range(4, sheet.max_row + 1):  # 假设数据从第4行开始
            original_value = sheet.cell(row, talent_code_col).value
            if not original_value:
                continue

            new_value = self.modify_value(str(original_value))
            sheet.cell(row, talent_code_col).value = new_value
            modified_count += 1
            print(f"  ✅ 修改字段：talent_code | 原值: {original_value} → 新值: {new_value}（行{row}）")

        print(f"  ✔️ 完成：共修改{modified_count}条talent_code记录")

    def process_delete_education_sheet(self, wb, sheet_name):
        print(f"处理 '{sheet_name}' 表单...")
        sheet = wb[sheet_name]
        rid_col = next((col for col in range(1, sheet.max_column + 1)
                        if sheet.cell(1, col).value == "rid"), None)
        if not rid_col:
            print(f"  ⚠️ 跳过：未找到'rid'列（第1行）")
            return

        sql = f"select rid from details.v_talent_education_latest limit {sheet.max_row - 1}"
        new_rids = [str(rid[0]) for rid in self.execute_query(sql)]

        for i, row in enumerate(range(2, sheet.max_row + 1)):
            if i < len(new_rids):
                sheet.cell(row, rid_col).value = new_rids[i]
            else:
                sheet.delete_rows(row, sheet.max_row - row + 1)
                break
        print(f"  ✔️ 完成：更新{len(new_rids)}条学历rid记录")

    def process_excel_file(self, file_path):
        print(f"\n处理文件: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path)
            valid_sheets = [s for s in wb.sheetnames if s in self.valid_sheets]
            if not valid_sheets:
                print(f"  ⚠️ 跳过：无有效表单")
                return

            print(f"  找到有效表单: {', '.join(valid_sheets)}")
            for sheet_name in valid_sheets:
                if sheet_name == "新增明细":
                    self.process_new_details_sheet(wb, sheet_name)
                elif sheet_name == "修改明细":
                    self.process_modify_details_sheet(wb, sheet_name)
                elif sheet_name == "删除明细":
                    self.process_delete_details_sheet(wb, sheet_name)
                elif sheet_name in ["新增学历", "修改学历", "人才现职"]:
                    self.process_talent_code_sheet(wb, sheet_name)
                elif sheet_name == "删除学历":
                    self.process_delete_education_sheet(wb, sheet_name)

            wb.save(file_path)
            print(f"  ✅ 保存：{file_path}")
        except Exception as e:
            print(f"  ❌ 错误：{str(e)[:50]}...（完整错误请查看堆栈）")

    def process_all_files(self):
        if not self.test_file_dir.exists():
            print(f"错误: 测试目录不存在 - {self.test_file_dir}")
            return
        if not self.test_file_dir.is_dir():
            print(f"错误: 非目录路径 - {self.test_file_dir}")
            return

        excel_files = list(self.test_file_dir.glob("*.xlsx")) + list(self.test_file_dir.glob("*.xls"))
        if not excel_files:
            print(f"提示: {self.test_file_dir}中无Excel文件")
            return

        print(f"开始处理 {len(excel_files)} 个文件（目录: {self.test_file_dir}）")
        for file in excel_files:
            self.process_excel_file(str(file))

    def run(self):  # 关键补充：明确存在的run方法
        print("开始Excel更新器...")
        self.process_all_files()
        print("\nExcel更新处理完成")


if __name__ == "__main__":
    updater = ExcelUpdater()
    updater.run()