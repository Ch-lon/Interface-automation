import openpyxl
from common.db_handler import db
from common.log_handler import logger




def get_cases_from_excel(file_address, sheet_name=None):
    """
    从Excel获取测试用例
    :param file_address: Excel文件路径
    :param sheet_name: sheet名
    :return: 列表
    """
    wb =openpyxl.load_workbook(file_address)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    # 定义标题字典
    title = {}
    for i in range(1, ws.max_column + 1):
        title[i] = ws.cell(row = 1, column = i).value
    # 定义结果列表（一行数据为一个元素）
    data = []
    for i in range(2, ws.max_row + 1):
        # 临时字典，存储每一行的数据
        temp = {}
        for j in range(1, ws.max_column + 1):
            temp[title[j]] = ws.cell(row = i, column = j).value
        data.append(temp)

    return data

def get_real_max_row(ws):
    """
    获取表格中的最大行（去除格式编辑/历史痕迹之后）
    :param ws: 工作表
    :return: 处理好的肉眼可见的最大行
    """
    # 从第一行开始倒序遍历
    for row in reversed(range(1, ws.max_row + 1)):
        # 检查该行是否有至少一个非空值（忽略仅格式单元格）
        if any(cell.value is not None for cell in ws[row]):
            return row
    return 0  # 空表

def db_result_handler(result : list):
    """
    对数据库查询结果进行格式处理
    :param result: 数据库查询结果(列表内嵌元组)，如[('elected_name',), ('elected_year',), ('talent_name',)]
    :return: 处理好的列表
    """
    ls = []
    for i in result:
        ls.append(i[0])
    return ls

def get_row_first_comment(file_path: str, sheet_name: str = None, row_num: int = 1) -> tuple:
    """
    获取某行的第一个批注的位置和内容
    :param file_path: 文件路径
    :param sheet_name: sheet名
    :param row_num: 行号
    :return: 返回格式：(单元格地址, 批注内容) 或 (None, None)
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        for cell in ws[row_num]:
            if cell.comment:
                cell_addr = f"{openpyxl.utils.get_column_letter(cell.column)}{cell.row}"
                wb.close()  # 找到后立即关闭文件
                return cell_addr, cell.comment.text.strip()

        wb.close()
        return None, "该行无批注"  # 明确无批注的情况

    except FileNotFoundError:
        return None, "错误：文件未找到"
    except Exception as e:
        wb.close()
        return None, f"错误：{str(e)}"

def get_row_all_comments(file_path: str, sheet_name: str = None, row_num: int = 1):
    """
    获取某一行的所有批注
    :param file_path: 文件路径
    :param sheet_name: sheet名
    :param row_num: 行号
    :return: 返回格式：{单元格地址: 批注内容}（如 {'B3': '重要数据', 'D3': '注意单位'}）
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        for cell in ws[row_num]:  # 遍历行中的每个单元格
            if cell.comment:
                cell_addr = f"{openpyxl.utils.get_column_letter(cell.column)}{cell.row}"
                result[cell_addr] = cell.comment.text.strip()

    except Exception as e:
        print(f"查找批注失败: {str(e)}")
    finally:
        wb.close()
    return result

def get_all_sheet_all_comments(file_path) -> dict:
    """
    获取所有有效sheet中的所有批注
    :param file_path: 文件路径
    :return: 结构化批注字典，格式：{sheet名-行号-列号: 批注内容}
    """
    VALID_SHEET_NAMES = {
        "新增明细", "修改明细", "删除明细",
        "人才现职", "新增学历", "修改学历", "删除学历"
    }
    result = {}
    try:
        # 移除read_only模式（批注只能在完整模式下读取）
        wb = openpyxl.load_workbook(file_path, data_only=True)  # 保留data_only不影响批注

        valid_sheets = [
            (name, wb[name])
            for name in wb.sheetnames
            if name in VALID_SHEET_NAMES
        ]

        for sheet_name, ws in valid_sheets:
            # 优化行遍历方式（直接使用行号循环，避免ReadOnlyCell问题）
            for row_num in range(1, ws.max_row + 1):
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.comment:
                        # 保留原键格式（sheet-行-列数字）
                        key = f"{sheet_name}-{row_num}-{col_num}"
                        # 处理富文本批注（兼容openpyxl 3.1+的Comment对象）
                        comment_text = (
                            cell.comment.text.strip()
                            if hasattr(cell.comment, 'text')
                            else str(cell.comment)
                        )
                        result[key] = comment_text

    except FileNotFoundError:
        print(f"错误：文件未找到 - {file_path}")
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"错误：非Excel文件或文件损坏 - {file_path}")
    except Exception as e:
        print(f"扫描批注失败：{str(e)}")
    finally:
        if 'wb' in locals():
            wb.close()  # 确保资源释放
    return result


if __name__ == '__main__':

    res = get_cases_from_excel(r'D:\Pycharm\Pycharm文件\AutoTest\test_data\cases.xlsx', 'Inspection')
    print(res)

