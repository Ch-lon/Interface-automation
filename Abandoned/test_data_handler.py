import openpyxl
from common.db_handler import db
from common.log_handler import logger




def get_cases_from_excel(file_address, sheet_name=None):
    """
    从Excel获取测试用例
    :param file_address: Excel文件路径
    :param sheet_name: sheet名
    :return:
    """
    wb =openpyxl.load_workbook(file_address)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    # 定义标题字典
    title = {}
    for i in range(1, ws.max_column + 1):
        title[i] = ws.cell(row = 1, column = i).value
    # 定义结果列表（一行数据为一个元素）
    data = []
    for i in range(2, ws.max_row + 1):
        # 临时字典，存储每一行的数据
        temp = {}
        for j in range(1, ws.max_column + 1):
            temp[title[j]] = ws.cell(row = i, column = j).value
        data.append(temp)

    return data

def get_real_max_row(ws):
    """
    获取表格中的最大行（去除格式编辑/历史痕迹之后）
    :param ws: 工作表
    :return: 处理好的肉眼可见的最大行
    """
    # 从第一行开始倒序遍历
    for row in reversed(range(1, ws.max_row + 1)):
        # 检查该行是否有至少一个非空值（忽略仅格式单元格）
        if any(cell.value is not None for cell in ws[row]):
            return row
    return 0  # 空表

def db_result_handler(result : list):
    """
    对数据库查询结果进行格式处理
    :param result: 数据库查询结果(列表内嵌元组)，如[('elected_name',), ('elected_year',), ('talent_name',)]
    :return: 处理好的列表
    """
    ls = []
    for i in result:
        ls.append(i[0])
    return ls

def judge_year(ls : list):
    """
    判断唯一校验字段列表中是否含有ver_no或elected_year
    :param ls: 被判断的唯一标识列表
    :return:
    """
    for i in ls:
        if i in ['ver_no', 'elected_year']:
            return i
    return False

def modify_cell(ws, row, column, year = True):
    """
    修改单元格的值
    :param ws: 工作表
    :param row: 单元格所在行号
    :param column: 单元格所在列号
    :param year: 默认修改的是年份的值
    :return:
    """
    try:
        if year:
            old_value = ws.cell(row = row, column = column).value
            new_value = int(old_value) + 1
            ws.cell(row=row, column=column).value = new_value
        else:
            n = 1
            old_value = ws.cell(row = row, column = column).value
            new_value = str(old_value) + str(n)
            ws.cell(row=row, column=column).value = new_value
    except Exception as e:
        raise e

def get_index(obj, s, start = 1, flag = 0):
    """
    获取索引和值
    :param obj: 对象
    :param s: 进行比对的字符串
    :param start: 索引开始的位置
    :param flag: 是否需要返回值，默认仅返回索引， 0代表仅需返回索引， 1代表需同时返回索引和值
    :return:
    """
    for index, cell in enumerate(obj, start = start):
        if cell.value == s:
            if flag == 0:
                return index
            else:
                return index, cell.value

def get_r_uf_index(modify_sheet, file_name, sheet_name = '修改明细'):
    """
    获取rid和update_fields的行号、列号
    :param modify_sheet: 修改sheet对象 worksheet
    :param file_name: 文件名
    :param sheet_name: sheet名
    :return: rid行号、rid列号、update_fields行号、update_fields列号
    """
    rid_row = 0
    rid_column = 0
    update_fields_row = 0
    update_fields_column = 0
    # 遍历前三行寻找rid和update_fields字段名位置
    for row_index in range(1, 4):
        for cell_column_index, cell in enumerate(modify_sheet[row_index], start=1):
            if cell.value == 'rid':
                rid_row = row_index
                rid_column = cell_column_index
                continue
            if cell.value == 'update_fields':
                update_fields_row = row_index
                update_fields_column = cell_column_index
        if rid_row and update_fields_row:
            break
    if not (rid_row or update_fields_row):
        raise f"{file_name}的{sheet_name}的前三行未找到rid和update_fields"
    return rid_row, rid_column, update_fields_row, update_fields_column

def get_excel_field(ws, row):
    """
    获取Excel中字段名
    :param ws: 工作表 worksheet
    :param row: 行号
    :return:
    """
    # 未剔除var_code/name
    result = []
    for i in ws[row]:
        result.append(i.value)
    return result

def judge_field_in_list(fields, ls):
    """
    判断一个列表中的字段是否在另一个列表中,并返回重合值
    :param fields: 待判断的字段列表
    :param ls: 列表
    :return:
    """
    result = []
    for i in fields:
        if i in ls:
            result.append(i)
    return result

def add_details_handler(wb, file_name):
    """
    处理新增明细sheet（字段值修改）
    :param wb: 工作簿 workbook
    :param file_name: Excel文件名
    :return:
    """
    add_details_sheet = wb['新增明细']
    # 定义var_code所处列数
    vc_column = get_index(add_details_sheet[3], 'var_code')
    max_row = get_real_max_row(add_details_sheet)
    # 遍历每一行
    for i in range(4, max_row + 1):
        # 定义var_code值变量
        var_Code_Value = add_details_sheet.cell(row=i, column=vc_column).value
        sql = f"select field_key from details.var_det_field where def_id = (select det_def_id from details.v_var where code = '{var_Code_Value}') and uni_check > 0"
        uni_check_list = db_result_handler(db.get_all(sql))
        year_name = judge_year(uni_check_list)
        if year_name:
            # 获取字段名所在列号
            column_name_index = get_index(add_details_sheet[3], year_name)
            # 修改字段
            modify_cell(add_details_sheet, i, column_name_index)
        else:
            raise f"文件{file_name}的第{i}行中的变量{var_Code_Value}的唯一标识中没有ver_no或elected_year"

def modify_details_handler(wb, file_name):
    """
    处理修改明细sheet（字段值修改）
    :param wb: 工作簿 workbook
    :param file_name: Excel文件名
    :return:
    """
    # 1. update_fields中填写ALL_FIELDS，需获取Excel中全部的字段名，需根据rid去数据库中查询到该条明细，之后获取查询结果中的var_code，然后查询唯一标识，接着再判断修改的字段中是否包含唯一标识，\
    # 如果包含唯一标识，则需对唯一标识进行修改；如果所修改字段中没有唯一标识，则无需更改
    # 2. 若update_fields中填写的是特定字段，则直接判断其中是否包含唯一标识，若包含则需修改，若不包含则无需修改
    modify_details_sheet = wb['修改明细']
    rid_row, rid_column, update_fields_row, update_fields_column = get_r_uf_index(modify_details_sheet, file_name)
    # 获取Excel中全部的字段名
    excel_fieldNames = get_excel_field(modify_details_sheet, update_fields_row)

    # 遍历每一行
    max_row = get_real_max_row(modify_details_sheet)
    for i in range(rid_row + 1, max_row + 1):
        rid_value = modify_details_sheet.cell(row = i, column = rid_column).value
        # Excel每一行的update_fields中填写的值
        ex_update_fields_value = modify_details_sheet.cell(row = i, column = update_fields_column).value

        sql = f'select field_key from details.var_det_field where uni_check > 0 and def_id = \
                (select det_def_id from details.variable where code = (select var_code from details.v_var_detail_xp where rid = {rid_value}))'
        # 数据库获取该变量的唯一标识
        db_uni_check_list = db_result_handler(db.get_all(sql))

        # 获取Excel中待修改字段中的唯一标识字段
        excel_uni_checks = judge_field_in_list(db_uni_check_list, excel_fieldNames)

        if ex_update_fields_value == 'ALL_FIELDS':
            if excel_uni_checks:
                # 若待修改字段中有唯一标识
                try:
                    field_column_index = get_index(modify_details_sheet[update_fields_row], excel_uni_checks[0])
                    modify_cell(modify_details_sheet, i, field_column_index, False)
                except Exception as e:
                    raise e
            else:
                # Excel字段中没有唯一标识（不做修改）
                logger.warning(f"文件{file_name}的修改sheet的第{i}行的待修改字段中没有唯一标识")
        else:
            # update_fields中填写的是特定字段
            res = judge_field_in_list(ex_update_fields_value, db_uni_check_list)
            if res:
                # update_fields中填写的有唯一标识
                field_column_index = get_index(modify_details_sheet[update_fields_row], res[0])
                modify_cell(modify_details_sheet, i, field_column_index, False)
            else:
                # update_fields中填写的没有唯一标识(不做修改)
                logger.warning(f"文件{file_name}的修改sheet的第{i}行的待修改字段中没有唯一标识")

def del_details_handler(wb, file_name):
    """
    处理删除明细sheet（修改rid）
    :param wb: 工作簿 workbook
    :param file_name: Excel文件名
    :return:
    """
    del_details_sheet = wb['删除明细']
    # 默认仅提供rid一列, 表头仅一行
    max_row = get_real_max_row(del_details_sheet)
    sql = f"select rid from details.v_var_detail_xp where var_code = 'VG0008' limit {max_row - 1}"
    new_rid_values = db_result_handler(db.get_all(sql))
    for i in range(2, max_row + 1):
        del_details_sheet.cell(row = i, column = 1).value = new_rid_values.pop()

def get_row_first_comment(file_path: str, sheet_name: str = None, row_num: int = 1) -> tuple:
    """
    获取某行的第一个批注的位置和内容
    :param file_path: 文件路径
    :param sheet_name: sheet名
    :param row_num: 行号
    :return: 返回格式：(单元格地址, 批注内容) 或 (None, None)
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        for cell in ws[row_num]:
            if cell.comment:
                cell_addr = f"{openpyxl.utils.get_column_letter(cell.column)}{cell.row}"
                wb.close()  # 找到后立即关闭文件
                return cell_addr, cell.comment.text.strip()

        wb.close()
        return None, "该行无批注"  # 明确无批注的情况

    except FileNotFoundError:
        return None, "错误：文件未找到"
    except Exception as e:
        wb.close()
        return None, f"错误：{str(e)}"

def get_row_all_comments(file_path: str, sheet_name: str = None, row_num: int = 1):
    """
    获取某一行的所有批注
    :param file_path: 文件路径
    :param sheet_name: sheet名
    :param row_num: 行号
    :return: 返回格式：{单元格地址: 批注内容}（如 {'B3': '重要数据', 'D3': '注意单位'}）
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        for cell in ws[row_num]:  # 遍历行中的每个单元格
            if cell.comment:
                cell_addr = f"{openpyxl.utils.get_column_letter(cell.column)}{cell.row}"
                result[cell_addr] = cell.comment.text.strip()

    except Exception as e:
        print(f"查找批注失败: {str(e)}")
    finally:
        wb.close()
    return result

def get_all_sheet_all_comments(file_path: str) -> dict:
    """
    获取所有有效sheet中的所有批注
    :param file_path: 文件路径
    :return: 结构化批注字典，格式：{sheet名-行号-列号: 批注内容}
    """
    VALID_SHEET_NAMES = {
        "新增明细", "修改明细", "删除明细",
        "人才现职", "新增学历", "修改学历", "删除学历"
    }
    result = {}
    try:
        # 移除read_only模式（批注只能在完整模式下读取）
        wb = openpyxl.load_workbook(file_path, data_only=True)  # 保留data_only不影响批注

        valid_sheets = [
            (name, wb[name])
            for name in wb.sheetnames
            if name in VALID_SHEET_NAMES
        ]

        for sheet_name, ws in valid_sheets:
            # 优化行遍历方式（直接使用行号循环，避免ReadOnlyCell问题）
            for row_num in range(1, ws.max_row + 1):
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if cell.comment:
                        # 保留原键格式（sheet-行-列数字）
                        key = f"{sheet_name}-{row_num}-{col_num}"
                        # 处理富文本批注（兼容openpyxl 3.1+的Comment对象）
                        comment_text = (
                            cell.comment.text.strip()
                            if hasattr(cell.comment, 'text')
                            else str(cell.comment)
                        )
                        result[key] = comment_text

    except FileNotFoundError:
        print(f"错误：文件未找到 - {file_path}")
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"错误：非Excel文件或文件损坏 - {file_path}")
    except Exception as e:
        print(f"扫描批注失败：{str(e)}")
    finally:
        if 'wb' in locals():
            wb.close()  # 确保资源释放
    return result


if __name__ == '__main__':

    res = get_cases_from_excel(r'D:\Pycharm\Pycharm文件\AutoTest\test_data\cases.xlsx', 'Inspection')
    print(res)

    # res = get_all_sheet_all_comments(r"C:\Users\user\Downloads\FeedBack_1744855927_复合上传.xlsx")
    # print(res)


    # res = get_row_all_comments(r'D:\Pycharm\Pycharm文件\AutoTest\test_data\error_file\FeedBack_1744686066_2.xlsx', '新增明细', 4)
    # print(res)



    # wb1 = openpyxl.load_workbook(r"C:\Users\user\Desktop\复合上传.xlsx")
    # del_details_handler(wb1, '复合上传.xlsx')
    # wb1.save(r"C:\Users\user\Desktop\复合上传.xlsx")


    # wb1 = openpyxl.load_workbook(r'D:\Pycharm\Pycharm文件\AutoTest\test_data\test_file\1.xlsx')
    # modify_details_handler(wb1, '1.xlsx')
    # wb1.save(r'D:\Pycharm\Pycharm文件\AutoTest\test_data\test_file\1.xlsx')

    # ws1 = wb1['修改明细']
    # res1 = get_excel_field(ws1, 3)
    # print(res1)

    # res1 = db_result_handler([('elected_name',), ('elected_year',), ('talent_name',)])
    # print(res1)
    # res2 = get_cases_from_excel(settings.TEST_DATA_DIR, sheet_name = 'upload')
    # print(res2)
    # wb1 = openpyxl.load_workbook(r'D:\Pycharm\Pycharm文件\AutoTest\test_data\test_file\1.xlsx')
    # add_details_handler(wb1, file_name='1.xlsx')
    # # 保存文件(直接覆盖原文件)
    # wb1.save(r'D:\Pycharm\Pycharm文件\AutoTest\test_data\test_file\1.xlsx')
